import os
import logging
from datetime import datetime, date, timedelta
from dataclasses import dataclass
from typing import List, Optional, Tuple

import openpyxl

logger = logging.getLogger("PlanMonitor")


@dataclass
class PlanRow:
    order_number: str
    machine_name: str
    production_date: date
    planned_qty: int


def find_latest_excel(folder: str) -> Optional[Tuple[str, datetime]]:
    """Trova il file Excel piu recente nella cartella specificata."""
    try:
        if not os.path.exists(folder):
            logger.error("Cartella planning non raggiungibile: %s", folder)
            return None
        if not os.path.isdir(folder):
            logger.error("Il percorso non e' una cartella: %s", folder)
            return None
    except OSError as e:
        logger.error("Errore accesso cartella planning: %s", e)
        return None

    excel_files = []
    try:
        for f in os.listdir(folder):
            if f.lower().endswith((".xlsx", ".xls")) and not f.startswith("~$"):
                full_path = os.path.join(folder, f)
                try:
                    mtime = os.path.getmtime(full_path)
                    excel_files.append((full_path, datetime.fromtimestamp(mtime)))
                except OSError:
                    continue
    except OSError as e:
        logger.error("Errore lettura cartella planning: %s", e)
        return None

    if not excel_files:
        logger.error("Nessun file Excel trovato in: %s", folder)
        return None

    excel_files.sort(key=lambda x: x[1], reverse=True)
    selected = excel_files[0]
    logger.info("File Excel selezionato: %s (modificato: %s)", selected[0], selected[1])
    return selected


def _parse_date_header(value) -> Optional[date]:
    """Converte un'intestazione colonna in data."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    logger.warning("Intestazione data non parsabile: '%s'", s)
    return None


def _parse_qty(value) -> int:
    """Converte un valore cella in quantita intera."""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value) if value == value else 0  # gestisce NaN
    s = str(value).strip()
    if not s:
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def parse_last_phase(file_path: str, sheet_name: str = "PlanningMachine") -> List[PlanRow]:
    """Parsa il foglio Last Phase e restituisce le righe del piano."""
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except Exception as e:
        logger.error("Errore apertura file Excel '%s': %s", file_path, e)
        return []

    if sheet_name not in wb.sheetnames:
        logger.error("Foglio '%s' non trovato nel file '%s'. Fogli: %s",
                      sheet_name, file_path, wb.sheetnames)
        wb.close()
        return []

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    wb.close()

    if not rows:
        logger.error("Foglio '%s' vuoto", sheet_name)
        return []

    header_row = rows[0]

    # Colonne da U in poi (indice 20+) contengono le date
    date_columns = {}  # indice colonna -> date
    for col_idx in range(20, len(header_row)):
        parsed = _parse_date_header(header_row[col_idx])
        if parsed:
            date_columns[col_idx] = parsed

    if not date_columns:
        logger.error("Nessuna data valida trovata nelle intestazioni da colonna U in poi")
        return []

    logger.info("Date trovate nelle intestazioni: %s", sorted(date_columns.values()))

    plan_rows = []
    for row_idx, row in enumerate(rows[1:], start=2):
        if len(row) < 21:
            continue

        # Colonna K (indice 10) = ordine
        raw_order = row[10]
        if raw_order is None:
            continue
        order_number = str(raw_order).lstrip("\u2022").strip()
        if not order_number:
            continue

        # Colonna E (indice 4) = macchina/fase
        raw_machine = row[4]
        if raw_machine is None:
            continue
        machine_name = str(raw_machine).strip()
        if not machine_name:
            continue

        # Quantita per ogni data
        for col_idx, prod_date in date_columns.items():
            if col_idx < len(row):
                qty = _parse_qty(row[col_idx])
                if qty > 0:
                    plan_rows.append(PlanRow(
                        order_number=order_number,
                        machine_name=machine_name,
                        production_date=prod_date,
                        planned_qty=qty
                    ))

    logger.info("Righe piano lette: %d (da %d righe dati)", len(plan_rows), len(rows) - 1)
    return plan_rows


def get_todays_plan(plan_rows: List[PlanRow], today: date = None) -> List[PlanRow]:
    """Filtra le righe del piano per la data odierna."""
    if today is None:
        today = date.today()
    result = [r for r in plan_rows if r.production_date == today]
    logger.info("Righe piano per oggi (%s): %d", today, len(result))
    return result


def _get_working_days_back(from_date: date, count: int) -> List[date]:
    """Restituisce le ultime 'count' giornate lavorative prima di from_date (escluso)."""
    days = []
    d = from_date - timedelta(days=1)
    while len(days) < count:
        if d.weekday() < 5:  # lun-ven
            days.append(d)
        d -= timedelta(days=1)
    return days


def _get_working_days_forward(from_date: date, count: int) -> List[date]:
    """Restituisce le prossime 'count' giornate lavorative dopo from_date (escluso)."""
    days = []
    d = from_date + timedelta(days=1)
    while len(days) < count:
        if d.weekday() < 5:
            days.append(d)
        d += timedelta(days=1)
    return days


def get_plan_for_dates(plan_rows: List[PlanRow], dates: List[date]) -> List[PlanRow]:
    """Filtra le righe del piano per un insieme di date."""
    date_set = set(dates)
    return [r for r in plan_rows if r.production_date in date_set]


def check_order_in_past_plan(all_plan: List[PlanRow], order_number: str,
                              today: date = None, lookback_days: int = 2) -> Optional[PlanRow]:
    """Cerca se l'ordine era pianificato nei giorni lavorativi precedenti.
    Se lookback_days e' None, cerca in tutti i giorni passati presenti nel piano."""
    if today is None:
        today = date.today()
    if lookback_days is None:
        # Cerca in tutte le date passate presenti nel piano, dalla piu' recente
        past_matches = [pr for pr in all_plan
                        if pr.order_number == order_number and pr.production_date < today]
        if past_matches:
            return max(past_matches, key=lambda pr: pr.production_date)
        return None
    past_days = _get_working_days_back(today, lookback_days)
    for d in past_days:
        for pr in all_plan:
            if pr.order_number == order_number and pr.production_date == d:
                return pr
    return None


def check_order_in_future_plan(all_plan: List[PlanRow], order_number: str,
                                today: date = None, lookahead_days: int = 3) -> Optional[PlanRow]:
    """Cerca se l'ordine e' pianificato nei prossimi giorni lavorativi (max lookahead_days)."""
    if today is None:
        today = date.today()
    future_days = _get_working_days_forward(today, lookahead_days)
    for d in future_days:
        for pr in all_plan:
            if pr.order_number == order_number and pr.production_date == d:
                return pr
    return None
